"""
excel_reader.py
---------------
读取 customer_master.xlsx，提供：
  - list_customers()       → 返回所有客户的简要列表
  - get_customer(id)       → 返回指定客户的完整数据（标准字段 dict）
  - list_field_dictionary() → 返回 Field Dictionary Sheet 的内容

Sheet 结构（按蓝图 §4 定义）：
  Sheet 1 "Customers"：客户主资料
    - customer_id, full_name, ic_no, date_of_birth, nationality, gender,
      marital_status, race, religion, mobile_no, home_tel, email,
      address_line1, address_line2, address_line3, postcode, city, state, country,
      employer_name, employer_address, monthly_income, annual_income,
      occupation, employment_type, years_with_employer,
      bank_name, bank_account_no, loan_amount, loan_tenure

  Sheet 2 "FieldDictionary"：字段同义词字典
    - standard_key, display_name, synonyms（逗号分隔）, data_type, example
"""

import os
from typing import Optional

import openpyxl

# 默认路径（相对于 backend/ 目录的父目录）
DEFAULT_XLSX_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "customer_master.xlsx"
)

# 标准字段列表（与 customer_master.xlsx Sheet1 的列名一一对应）
CUSTOMER_STANDARD_FIELDS = [
    "customer_id",
    "full_name",
    "ic_no",
    "date_of_birth",
    "nationality",
    "gender",
    "marital_status",
    "race",
    "religion",
    "mobile_no",
    "home_tel",
    "email",
    "address_line1",
    "address_line2",
    "address_line3",
    "postcode",
    "city",
    "state",
    "country",
    "employer_name",
    "employer_address",
    "monthly_income",
    "annual_income",
    "occupation",
    "employment_type",
    "years_with_employer",
    "bank_name",
    "bank_account_no",
    "loan_amount",
    "loan_tenure",
]


def _get_xlsx_path(xlsx_path: Optional[str] = None) -> str:
    return xlsx_path or DEFAULT_XLSX_PATH


def list_customers(xlsx_path: Optional[str] = None) -> list[dict]:
    """
    返回所有客户的简要列表。
    每条记录只包含：customer_id, full_name, ic_no, mobile_no
    """
    path = _get_xlsx_path(xlsx_path)
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _get_sheet(wb, "Customers")
    if ws is None:
        wb.close()
        return []

    headers = _read_headers(ws)
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        result.append({
            "customer_id": record.get("customer_id"),
            "full_name": record.get("full_name", ""),
            "ic_no": record.get("ic_no", ""),
            "mobile_no": record.get("mobile_no", ""),
        })

    wb.close()
    return [r for r in result if r["customer_id"] is not None]


def get_customer(customer_id, xlsx_path: Optional[str] = None) -> Optional[dict]:
    """
    返回指定 customer_id 的完整客户数据。
    键格式：customer.<field_name>，与 field_normalizer 的标准键一致。
    返回 None 表示找不到该客户。
    """
    path = _get_xlsx_path(xlsx_path)
    if not os.path.exists(path):
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _get_sheet(wb, "Customers")
    if ws is None:
        wb.close()
        return None

    headers = _read_headers(ws)
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        cid = record.get("customer_id")
        # 支持 int/str 两种 customer_id 比较
        if str(cid) == str(customer_id):
            wb.close()
            # 转换为 customer.xxx 格式键
            return _to_standard_dict(record)

    wb.close()
    return None


def _to_standard_dict(record: dict) -> dict:
    """
    将 Excel 行转换为标准字段 dict：
    {"full_name": "张三"} → {"customer.full_name": "张三", ...}
    同时保留原始键以兼容两种访问方式。
    """
    result = {}
    for k, v in record.items():
        # 标准键（带前缀）
        result[f"customer.{k}"] = _format_value(v)
        # 裸键（无前缀，部分模块可能直接用）
        result[k] = _format_value(v)
    return result


def _format_value(value) -> str:
    """将 Excel 单元格值转为字符串，处理 None 和日期"""
    if value is None:
        return ""
    # openpyxl 的日期类型
    try:
        from datetime import date, datetime
        if isinstance(value, (date, datetime)):
            return value.strftime("%d/%m/%Y")
    except ImportError:
        pass
    return str(value).strip()


def list_field_dictionary(xlsx_path: Optional[str] = None) -> list[dict]:
    """
    读取 FieldDictionary Sheet，返回字段字典列表。
    供 field_normalizer 动态加载用户自定义同义词。
    """
    path = _get_xlsx_path(xlsx_path)
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _get_sheet(wb, "FieldDictionary")
    if ws is None:
        wb.close()
        return []

    headers = _read_headers(ws)
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get("standard_key"):
            result.append(record)

    wb.close()
    return result


def _get_sheet(wb: openpyxl.Workbook, name: str):
    """不区分大小写地获取 Sheet"""
    for title in wb.sheetnames:
        if title.strip().lower() == name.lower():
            return wb[title]
    # 退而求其次：按顺序返回第一个 Sheet（若 name == "Customers"）
    if name.lower() == "customers" and wb.sheetnames:
        return wb[wb.sheetnames[0]]
    return None


def _read_headers(ws) -> list[str]:
    """读取 Sheet 第一行作为列名（小写、去空格）"""
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
        if cell is not None:
            headers.append(str(cell).strip().lower().replace(" ", "_"))
        else:
            headers.append("")
    return headers
