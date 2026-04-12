"""
customer_store.py
-----------------
customers 表的 CRUD 层。
数据库主数据源，替代 customer_master.xlsx 直读。

公开函数：
  list_customers(page, page_size, q)  → 分页查询
  get_customer_by_id(customer_id)     → 单个客户（用于填表，带 customer.xxx 键）
  create_customer(data)               → 新建
  update_customer(id, data)           → 更新
  delete_customer(id)                 → 删除
  bulk_customers(create, update, delete) → 批量
  import_from_xlsx(file_bytes)        → 从 Excel 导入
  export_to_xlsx()                    → 导出为 Excel bytes
"""

import sqlite3
import logging
from datetime import datetime
from typing import Optional
from pathlib import Path

logger = logging.getLogger("customer_store")

# DB 路径与 template_store 保持一致
_DB_PATH = Path(__file__).parent.parent / "database" / "templates.db"

# 客户表的所有可写字段（id / created_at / updated_at 除外）
CUSTOMER_FIELDS = [
    "customer_id",
    "full_name",
    "ic_no",
    "date_of_birth",
    "nationality",
    "gender",
    "marital_status",
    "race",
    "religion",
    "mobile_no",
    "home_tel",
    "email",
    "address_line1",
    "address_line2",
    "address_line3",
    "postcode",
    "city",
    "state",
    "country",
    "employer_name",
    "employer_address",
    "monthly_income",
    "annual_income",
    "occupation",
    "employment_type",
    "years_with_employer",
    "bank_name",
    "bank_account_no",
    "loan_amount",
    "loan_tenure",
]

REQUIRED_FIELDS = {"customer_id", "full_name"}


def _conn():
    con = sqlite3.connect(str(_DB_PATH))
    con.row_factory = sqlite3.Row
    return con


def _row_to_dict(row) -> dict:
    return dict(row)


# ─────────────────────────────────────────────────────────────
#  查询
# ─────────────────────────────────────────────────────────────

def list_customers(
    page: int = 1,
    page_size: int = 50,
    q: Optional[str] = None,
) -> dict:
    """
    分页查询客户列表。
    q: 搜索关键字，匹配 full_name / customer_id / ic_no / mobile_no
    """
    offset = (page - 1) * page_size
    with _conn() as con:
        if q:
            like = f"%{q}%"
            where = (
                "WHERE full_name LIKE ? OR customer_id LIKE ? "
                "OR ic_no LIKE ? OR mobile_no LIKE ?"
            )
            params_count = (like, like, like, like)
            params_data  = (like, like, like, like, page_size, offset)
        else:
            where = ""
            params_count = ()
            params_data  = (page_size, offset)

        total = con.execute(
            f"SELECT COUNT(*) FROM customers {where}", params_count
        ).fetchone()[0]

        rows = con.execute(
            f"SELECT * FROM customers {where} ORDER BY id LIMIT ? OFFSET ?",
            params_data,
        ).fetchall()

    return {
        "items"    : [_row_to_dict(r) for r in rows],
        "total"    : total,
        "page"     : page,
        "page_size": page_size,
    }


def get_customer_by_cid(customer_id: str) -> Optional[dict]:
    """
    通过 customer_id（如 C001）查询单个客户，
    返回带 customer.xxx 前缀的标准字段 dict（供填表模块使用）。
    """
    with _conn() as con:
        row = con.execute(
            "SELECT * FROM customers WHERE customer_id = ?", (str(customer_id),)
        ).fetchone()
    if row is None:
        return None
    record = _row_to_dict(row)
    return _to_standard_dict(record)


def get_customer_raw(record_id: int) -> Optional[dict]:
    """通过数据库主键查询，返回原始 dict（供管理接口使用）。"""
    with _conn() as con:
        row = con.execute(
            "SELECT * FROM customers WHERE id = ?", (record_id,)
        ).fetchone()
    return _row_to_dict(row) if row else None


# ─────────────────────────────────────────────────────────────
#  创建
# ─────────────────────────────────────────────────────────────

def create_customer(data: dict) -> dict:
    """
    新建客户。必须包含 customer_id / full_name。
    返回新建后的完整记录。
    """
    _validate_required(data)
    cols = [f for f in CUSTOMER_FIELDS if f in data]
    placeholders = ", ".join("?" * len(cols))
    col_names    = ", ".join(cols)
    values       = [str(data[c]) if data[c] is not None else None for c in cols]

    with _conn() as con:
        cur = con.execute(
            f"INSERT INTO customers ({col_names}) VALUES ({placeholders})",
            values,
        )
        con.commit()
        new_id = cur.lastrowid

    return get_customer_raw(new_id)


# ─────────────────────────────────────────────────────────────
#  更新
# ─────────────────────────────────────────────────────────────

def update_customer(record_id: int, data: dict) -> Optional[dict]:
    """
    更新客户（按数据库 id）。data 只需包含要修改的字段。
    返回更新后的完整记录，或 None 若 id 不存在。
    """
    # 只允许更新 CUSTOMER_FIELDS 中的列
    cols = [f for f in CUSTOMER_FIELDS if f in data]
    if not cols:
        return get_customer_raw(record_id)

    set_clause = ", ".join(f"{c} = ?" for c in cols)
    set_clause += ", updated_at = ?"
    values = [str(data[c]) if data[c] is not None else None for c in cols]
    values.append(datetime.utcnow().isoformat())

    with _conn() as con:
        con.execute(
            f"UPDATE customers SET {set_clause} WHERE id = ?",
            [*values, record_id],
        )
        con.commit()

    return get_customer_raw(record_id)


# ─────────────────────────────────────────────────────────────
#  删除
# ─────────────────────────────────────────────────────────────

def delete_customer(record_id: int) -> bool:
    """删除指定 id 的客户。返回 True 表示成功删除。"""
    with _conn() as con:
        cur = con.execute("DELETE FROM customers WHERE id = ?", (record_id,))
        con.commit()
    return cur.rowcount > 0


# ─────────────────────────────────────────────────────────────
#  批量操作
# ─────────────────────────────────────────────────────────────

def bulk_customers(
    create_list: list[dict],
    update_list: list[dict],
    delete_ids:  list[int],
) -> dict:
    """
    一次性处理所有新增 / 修改 / 删除。
    返回摘要 {created, updated, deleted, errors}。
    """
    created = 0
    updated = 0
    deleted = 0
    errors  = []

    for item in create_list:
        try:
            create_customer(item)
            created += 1
        except Exception as e:
            errors.append({"action": "create", "data": item, "error": str(e)})

    for item in update_list:
        rid = item.get("id")
        if not rid:
            errors.append({"action": "update", "data": item, "error": "missing id"})
            continue
        try:
            update_customer(int(rid), {k: v for k, v in item.items() if k != "id"})
            updated += 1
        except Exception as e:
            errors.append({"action": "update", "id": rid, "error": str(e)})

    for rid in delete_ids:
        try:
            delete_customer(int(rid))
            deleted += 1
        except Exception as e:
            errors.append({"action": "delete", "id": rid, "error": str(e)})

    return {
        "created": created,
        "updated": updated,
        "deleted": deleted,
        "errors" : errors,
    }


# ─────────────────────────────────────────────────────────────
#  Excel 导入 / 导出
# ─────────────────────────────────────────────────────────────

# Excel 列名 → DB 列名映射（大小写不敏感，空格→下划线）
_XLSX_COL_ALIAS = {
    "customerid": "customer_id",
    "customer_id": "customer_id",
    "fullname": "full_name",
    "full_name": "full_name",
    "icno": "ic_no",
    "ic_no": "ic_no",
    "nric": "ic_no",
    "mykad": "ic_no",
    "mobileno": "mobile_no",
    "mobile_no": "mobile_no",
    "phone": "mobile_no",
    "tel": "mobile_no",
    "hometel": "home_tel",
    "home_tel": "home_tel",
    "dateofbirth": "date_of_birth",
    "date_of_birth": "date_of_birth",
    "dob": "date_of_birth",
    "addressline1": "address_line1",
    "address_line1": "address_line1",
    "addressline2": "address_line2",
    "address_line2": "address_line2",
    "addressline3": "address_line3",
    "address_line3": "address_line3",
    "yearswithemployer": "years_with_employer",
    "years_with_employer": "years_with_employer",
    "employmenttype": "employment_type",
    "employment_type": "employment_type",
    "employername": "employer_name",
    "employer_name": "employer_name",
    "employeraddress": "employer_address",
    "employer_address": "employer_address",
    "monthlyincome": "monthly_income",
    "monthly_income": "monthly_income",
    "annualincome": "annual_income",
    "annual_income": "annual_income",
    "bankname": "bank_name",
    "bank_name": "bank_name",
    "bankaccountno": "bank_account_no",
    "bank_account_no": "bank_account_no",
    "loanamount": "loan_amount",
    "loan_amount": "loan_amount",
    "loantenure": "loan_tenure",
    "loan_tenure": "loan_tenure",
    "maritalstatus": "marital_status",
    "marital_status": "marital_status",
}


def _normalize_col(name: str) -> str:
    """将 Excel 列名标准化为 DB 列名。"""
    key = name.strip().lower().replace(" ", "_").replace("-", "_")
    return _XLSX_COL_ALIAS.get(key, key)


def import_from_xlsx(file_bytes: bytes, mode: str = "upsert") -> dict:
    """
    解析 Excel 并导入 customers 表。
    mode:
      - "upsert"  : 按 customer_id 做 upsert（默认）
      - "replace" : 清空表后全量导入
    返回 {imported, skipped, errors}
    """
    import io
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl 未安装，无法导入 Excel")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # 找 Customers sheet（不区分大小写）
    ws = None
    for title in wb.sheetnames:
        if title.strip().lower() in ("customers", "customer"):
            ws = wb[title]
            break
    if ws is None and wb.sheetnames:
        ws = wb[wb.sheetnames[0]]
    if ws is None:
        raise ValueError("Excel 中找不到 Customers Sheet")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        raise ValueError("Excel 文件为空")

    headers = [_normalize_col(str(h)) if h else "" for h in raw_headers]

    data_rows = []
    for row in rows_iter:
        record = {}
        for h, v in zip(headers, row):
            if h and h in CUSTOMER_FIELDS:
                record[h] = "" if v is None else str(v).strip()
        if record.get("customer_id") or record.get("full_name"):
            data_rows.append(record)

    wb.close()

    imported = 0
    skipped  = 0
    errors   = []

    if mode == "replace":
        with _conn() as con:
            con.execute("DELETE FROM customers")
            con.commit()

    for record in data_rows:
        if not record.get("customer_id") or not record.get("full_name"):
            skipped += 1
            continue
        try:
            if mode == "upsert":
                # 检查是否已存在
                with _conn() as con:
                    existing = con.execute(
                        "SELECT id FROM customers WHERE customer_id = ?",
                        (record["customer_id"],),
                    ).fetchone()
                if existing:
                    update_customer(existing["id"], record)
                else:
                    create_customer(record)
            else:
                create_customer(record)
            imported += 1
        except Exception as e:
            errors.append({"customer_id": record.get("customer_id"), "error": str(e)})

    return {"imported": imported, "skipped": skipped, "errors": errors}


def export_to_xlsx() -> bytes:
    """
    从 customers 表导出为 Excel bytes。
    列顺序与 CUSTOMER_FIELDS 相同（去掉内部 id）。
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl 未安装，无法导出 Excel")

    with _conn() as con:
        rows = con.execute(
            "SELECT * FROM customers ORDER BY id"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # 表头（使用 CUSTOMER_FIELDS 顺序，首字母大写友好显示）
    display_headers = [f.replace("_", " ").title() for f in CUSTOMER_FIELDS]
    ws.append(display_headers)

    # 表头样式
    header_fill = PatternFill("solid", fgColor="4A3F6B")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # 数据行
    for row in rows:
        d = _row_to_dict(row)
        ws.append([d.get(f, "") or "" for f in CUSTOMER_FIELDS])

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
#  辅助
# ─────────────────────────────────────────────────────────────

def _validate_required(data: dict):
    missing = [f for f in REQUIRED_FIELDS if not data.get(f)]
    if missing:
        raise ValueError(f"缺少必填字段：{', '.join(missing)}")


def _to_standard_dict(record: dict) -> dict:
    """
    将 DB 行转换为标准字段 dict：
    {"full_name": "张三"} → {"customer.full_name": "张三", "full_name": "张三", ...}
    供填表模块使用。
    """
    result = {}
    for k, v in record.items():
        v_str = "" if v is None else str(v)
        result[f"customer.{k}"] = v_str
        result[k] = v_str
    return result
