"""
generate_excel.py
-----------------
生成 customer_master.xlsx，包含两个 Sheet：
  1. Customers    — 客户主资料（含示例数据）
  2. FieldDictionary — 字段同义词字典

运行：python generate_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
#  列定义（与 excel_reader.py CUSTOMER_STANDARD_FIELDS 一致）
# ──────────────────────────────────────────────────────────────

CUSTOMER_COLUMNS = [
    ("customer_id",           "客户编号",     "ID",         10),
    ("full_name",             "姓名",         "Text",       25),
    ("ic_no",                 "身份证号码",   "Text",       20),
    ("date_of_birth",         "出生日期",     "Date",       15),
    ("nationality",           "国籍",         "Text",       15),
    ("gender",                "性别",         "Text",       10),
    ("marital_status",        "婚姻状况",     "Text",       15),
    ("race",                  "种族",         "Text",       12),
    ("religion",              "宗教",         "Text",       12),
    ("mobile_no",             "手机号码",     "Text",       18),
    ("home_tel",              "家庭电话",     "Text",       18),
    ("email",                 "电子邮件",     "Text",       28),
    ("address_line1",         "地址第一行",   "Text",       40),
    ("address_line2",         "地址第二行",   "Text",       40),
    ("address_line3",         "地址第三行",   "Text",       40),
    ("postcode",              "邮政编码",     "Text",       10),
    ("city",                  "城市",         "Text",       15),
    ("state",                 "州属",         "Text",       15),
    ("country",               "国家",         "Text",       15),
    ("employer_name",         "雇主名称",     "Text",       30),
    ("employer_address",      "工作地址",     "Text",       40),
    ("monthly_income",        "月收入",       "Number",     15),
    ("annual_income",         "年收入",       "Number",     15),
    ("occupation",            "职业",         "Text",       20),
    ("employment_type",       "雇佣类型",     "Text",       20),
    ("years_with_employer",   "工龄(年)",     "Number",     12),
    ("bank_name",             "银行名称",     "Text",       20),
    ("bank_account_no",       "银行账号",     "Text",       22),
    ("loan_amount",           "贷款金额",     "Number",     15),
    ("loan_tenure",           "贷款期限(月)", "Number",     12),
]

# 示例客户数据
SAMPLE_CUSTOMERS = [
    [
        "C001", "CHAN MEI LING", "880515-14-5678", "15/05/1988",
        "Malaysian", "Female", "Married", "Chinese", "Buddhism",
        "012-3456789", "03-12345678", "chanmeiling@email.com",
        "No. 12, Jalan Bukit Bintang", "Kuala Lumpur", "", "50200",
        "Kuala Lumpur", "Wilayah Persekutuan", "Malaysia",
        "ABC Technology Sdn Bhd", "Level 10, Menara ABC, KLCC, KL",
        8500, 102000, "Software Engineer", "Full-time", 5,
        "Maybank", "1234567890", 200000, 360,
    ],
    [
        "C002", "AHMAD BIN RAHMAN", "791230-10-2345", "30/12/1979",
        "Malaysian", "Male", "Married", "Malay", "Islam",
        "016-9876543", "03-98765432", "ahmad.rahman@email.com",
        "No. 45, Taman Maju", "Petaling Jaya", "Selangor", "47810",
        "Petaling Jaya", "Selangor", "Malaysia",
        "XYZ Manufacturing Sdn Bhd", "Lot 5, Jalan Industri, Shah Alam",
        6200, 74400, "Production Manager", "Full-time", 12,
        "CIMB Bank", "9876543210", 150000, 240,
    ],
    [
        "C003", "MUTHU KRISHNAN A/L SUPPIAH", "920318-07-8901", "18/03/1992",
        "Malaysian", "Male", "Single", "Indian", "Hindu",
        "011-22334455", "", "muthu.k@email.com",
        "No. 7, Jalan Indah 3", "Subang Jaya", "Selangor", "47500",
        "Subang Jaya", "Selangor", "Malaysia",
        "DEF Consultancy Sdn Bhd", "Plaza Masalam, Shah Alam",
        5800, 69600, "Business Analyst", "Full-time", 3,
        "Public Bank", "5432109876", 100000, 180,
    ],
    [
        "C004", "LIM SOO PING", "850722-12-3456", "22/07/1985",
        "Malaysian", "Female", "Married", "Chinese", "Taoism",
        "013-7654321", "04-1234567", "lim.sooping@email.com",
        "No. 23, Jalan Masjid", "George Town", "Pulau Pinang", "10050",
        "George Town", "Pulau Pinang", "Malaysia",
        "GHI Trading Sdn Bhd", "Penang Sentral, Butterworth",
        7000, 84000, "Sales Manager", "Full-time", 8,
        "Hong Leong Bank", "3210987654", 180000, 300,
    ],
    [
        "C005", "NUR FARAHANA BINTI ZULKIFLI", "950601-08-7890", "01/06/1995",
        "Malaysian", "Female", "Single", "Malay", "Islam",
        "017-1122334", "", "farahana.z@email.com",
        "A-12-5, Residensi Vista Komanwel", "Bukit Jalil", "Kuala Lumpur", "57000",
        "Kuala Lumpur", "Wilayah Persekutuan", "Malaysia",
        "JKL Healthcare Sdn Bhd", "Hospital Pantai, Bangsar",
        4500, 54000, "Nurse", "Full-time", 2,
        "RHB Bank", "6789012345", 80000, 120,
    ],
]

# ──────────────────────────────────────────────────────────────
#  字段字典 Sheet 数据
# ──────────────────────────────────────────────────────────────

FIELD_DICTIONARY = [
    ("customer.full_name", "Full Name / 姓名",
     "Full Name,Name,Nama,Nama Penuh,Applicant Name,姓名,申请人姓名",
     "Text", "CHAN MEI LING"),
    ("customer.ic_no", "IC No / 身份证号码",
     "IC No,NRIC No,MyKad No,No KP,Identity Card No,身份证号码",
     "Text", "880515-14-5678"),
    ("customer.date_of_birth", "Date of Birth / 出生日期",
     "Date of Birth,DOB,Tarikh Lahir,出生日期",
     "Date", "15/05/1988"),
    ("customer.nationality", "Nationality / 国籍",
     "Nationality,Warganegara,国籍",
     "Text", "Malaysian"),
    ("customer.gender", "Gender / 性别",
     "Gender,Sex,Jantina,性别",
     "Text", "Female"),
    ("customer.marital_status", "Marital Status / 婚姻状况",
     "Marital Status,Status Perkahwinan,婚姻状况",
     "Text", "Married"),
    ("customer.mobile_no", "Mobile No / 手机号码",
     "Mobile No,Phone,Tel No,No Tel Bimbit,手机号码",
     "Text", "012-3456789"),
    ("customer.email", "Email / 电邮",
     "Email,E-mail,电子邮件",
     "Text", "chanmeiling@email.com"),
    ("customer.address_line1", "Address / 地址",
     "Address,Home Address,Alamat,地址",
     "Text", "No. 12, Jalan Bukit Bintang"),
    ("customer.postcode", "Postcode / 邮编",
     "Postcode,Postal Code,Poskod,邮政编码",
     "Text", "50200"),
    ("customer.city", "City / 城市",
     "City,Town,Bandar,城市",
     "Text", "Kuala Lumpur"),
    ("customer.state", "State / 州属",
     "State,Negeri,州属",
     "Text", "Wilayah Persekutuan"),
    ("customer.employer_name", "Employer Name / 雇主",
     "Employer Name,Company Name,Nama Majikan,雇主名称",
     "Text", "ABC Technology Sdn Bhd"),
    ("customer.monthly_income", "Monthly Income / 月收入",
     "Monthly Income,Gross Monthly Salary,Pendapatan Bulanan,月收入",
     "Number", "8500"),
    ("customer.occupation", "Occupation / 职业",
     "Occupation,Job Title,Pekerjaan,职业",
     "Text", "Software Engineer"),
]


# ──────────────────────────────────────────────────────────────
#  样式工具函数
# ──────────────────────────────────────────────────────────────

def header_style():
    """表头样式：深紫色背景，白色加粗文字"""
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    fill = PatternFill("solid", fgColor="4A3F6B")   # Morandi Violet
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, alignment


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_header_row(ws, headers: list[str], col_widths: list[int]):
    """写入表头并应用样式"""
    font, fill, alignment = header_style()
    border = thin_border()
    ws.row_dimensions[1].height = 30

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_data_row(ws, row_idx: int, values: list, col_count: int):
    """写入数据行并应用样式"""
    border = thin_border()
    alt_fill = PatternFill("solid", fgColor="F5F3F8") if row_idx % 2 == 0 else None
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if alt_fill:
            cell.fill = alt_fill


# ──────────────────────────────────────────────────────────────
#  主函数
# ──────────────────────────────────────────────────────────────

def generate_excel(output_path: str = "customer_master.xlsx"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Customers ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Customers"
    ws1.freeze_panes = "A2"     # 冻结表头

    headers_en = [col[0] for col in CUSTOMER_COLUMNS]
    headers_cn = [col[1] for col in CUSTOMER_COLUMNS]
    widths = [col[3] for col in CUSTOMER_COLUMNS]

    # 第 1 行：英文字段名
    apply_header_row(ws1, headers_en, widths)

    # 写入示例数据
    for row_idx, row_data in enumerate(SAMPLE_CUSTOMERS, start=2):
        apply_data_row(ws1, row_idx, row_data, len(CUSTOMER_COLUMNS))

    # ── Sheet 2: FieldDictionary ──────────────────────────────
    ws2 = wb.create_sheet("FieldDictionary")
    ws2.freeze_panes = "A2"
    fd_headers = ["standard_key", "display_name", "synonyms", "data_type", "example"]
    fd_widths = [30, 25, 60, 12, 25]
    apply_header_row(ws2, fd_headers, fd_widths)

    for row_idx, row_data in enumerate(FIELD_DICTIONARY, start=2):
        apply_data_row(ws2, row_idx, list(row_data), len(fd_headers))

    # ── Sheet 3: Instructions ─────────────────────────────────
    ws3 = wb.create_sheet("使用说明")
    instructions = [
        ["智能表格自动填写系统 — customer_master.xlsx 使用说明"],
        [""],
        ["一、Customers Sheet"],
        ["  • 每一行代表一位客户的完整资料"],
        ["  • customer_id 为唯一标识符（不可重复）"],
        ["  • date_of_birth 请使用 DD/MM/YYYY 格式"],
        ["  • monthly_income / annual_income 请填入数字（不含货币符号）"],
        [""],
        ["二、FieldDictionary Sheet"],
        ["  • standard_key：系统内部使用的字段键（请勿修改）"],
        ["  • synonyms：该字段的所有中英文别名，用英文逗号分隔"],
        ["  • 可以添加新的同义词，系统重启后自动生效"],
        [""],
        ["三、注意事项"],
        ["  • 不要修改第一行（列名）"],
        ["  • 新增客户时请确保 customer_id 唯一"],
        ["  • 文件保存后，系统会自动读取最新数据（无需重启）"],
    ]
    ws3.column_dimensions["A"].width = 70
    for row_idx, row in enumerate(instructions, start=1):
        cell = ws3.cell(row=row_idx, column=1, value=row[0] if row else "")
        if row_idx == 1:
            cell.font = Font(bold=True, size=13, color="4A3F6B")
        elif row[0].startswith("一") or row[0].startswith("二") or row[0].startswith("三"):
            cell.font = Font(bold=True, size=11)

    wb.save(output_path)
    print(f"✓ customer_master.xlsx 已生成：{output_path}")
    print(f"  Customers Sheet：{len(SAMPLE_CUSTOMERS)} 条示例记录")
    print(f"  FieldDictionary Sheet：{len(FIELD_DICTIONARY)} 个字段定义")


if __name__ == "__main__":
    import sys
    output = sys.argv[1] if len(sys.argv) > 1 else "customer_master.xlsx"
    generate_excel(output)
